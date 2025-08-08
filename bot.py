import os
import io
import json
import threading
from datetime import datetime

from flask import Flask, request, jsonify
from telegram import Update, InputFile
from telegram.ext import Application, CommandHandler, MessageHandler, ContextTypes, filters

import stripe

# =============== Config ===============
TOKEN = os.getenv("BOT_TOKEN")
PUBLIC_BASE_URL = os.getenv("PUBLIC_BASE_URL", "https://example.com")

STRIPE_PUBLISHABLE_KEY = os.getenv("STRIPE_PUBLISHABLE_KEY")
STRIPE_SECRET_KEY = os.getenv("STRIPE_SECRET_KEY")
STRIPE_WEBHOOK_SECRET = os.getenv("STRIPE_WEBHOOK_SECRET")

stripe.api_key = STRIPE_SECRET_KEY

# =============== App State ===============
# STORE[user_id][period] = [ {amount, category, notes, ts}, ... ]
STORE = {}
MODE = {}                 # user_id -> "add" or None
CURRENT_PERIOD = {}       # user_id -> "YYYY-MM"
UNLOCKED = {}             # UNLOCKED[user_id] = set({"YYYY-MM", ...})

def _this_month() -> str:
    return datetime.utcnow().strftime("%Y-%m")

def _ensure_user_period(user_id: int) -> str:
    period = CURRENT_PERIOD.get(user_id)
    if not period:
        period = _this_month()
        CURRENT_PERIOD[user_id] = period
    STORE.setdefault(user_id, {}).setdefault(period, [])
    return period

def add_expense_to_store(user_id: int, amount: float, category: str, notes: str):
    period = _ensure_user_period(user_id)
    STORE[user_id][period].append({
        "amount": amount,
        "category": category,
        "notes": notes,
        "ts": datetime.utcnow()
    })

def clear_current_month(user_id: int):
    period = _ensure_user_period(user_id)
    STORE[user_id][period] = []

def get_current_month_expenses(user_id: int):
    period = _ensure_user_period(user_id)
    return STORE[user_id][period]

def _is_unlocked(user_id: int) -> bool:
    period = CURRENT_PERIOD.get(user_id, _this_month())
    return period in UNLOCKED.get(user_id, set())

def _mark_unlocked(user_id: int, period: str):
    UNLOCKED.setdefault(user_id, set()).add(period)

# =============== Parsing helpers ===============
def parse_free_expense(text: str):
    """
    Accepts a single line like:
      "1200 rent"
      "$50 groceries milk"
    Returns (amount, category, notes) or None.
    """
    parts = text.strip().split()
    if not parts:
        return None
    first = parts[0].replace("$", "").replace(",", "")
    try:
        amt = float(first)
    except ValueError:
        return None
    tokens = parts[1:]
    stopwords = {"for", "on", "to", "the", "a", "an", "my"}
    while tokens and tokens[0].lower() in stopwords:
        tokens = tokens[1:]
    cat = tokens[0] if tokens else "uncategorized"
    notes = " ".join(tokens[1:]) if len(tokens) > 1 else ""
    return amt, cat, notes

def parse_expense_block(text: str):
    """
    Accept multiple expenses separated by newlines OR slashes.
    Examples:
      1200 rent
      500 food
      200 insurance
    or:
      1200 rent / 500 food / 200 insurance
    """
    if ("\n" not in text) and ("/" in text):
        lines = [p.strip() for p in text.split("/") if p.strip()]
    else:
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    items, errors = [], []
    for raw in lines:
        parsed = parse_free_expense(raw)
        if parsed:
            items.append(parsed)
        else:
            errors.append(raw)
    return items, errors

# =============== Telegram Commands ===============
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    CURRENT_PERIOD[user_id] = _this_month()
    MODE[user_id] = "add"
    _ensure_user_period(user_id)
    await update.message.reply_text(
        "Welcome to Budget Wizard 🧙‍♂️\n"
        "Enter **monthly** expenses like:\n"
        "• 1200 rent\n• 60 phone\n• 230 car_insurance\n\n"
        "Paste multiple lines or use slashes: 1200 rent / 500 food / 200 insurance\n"
        "Shortcuts: **view**, **generate**, **export**, **unlock**, **done**.\n"
        "Type **reset** to start a new month."
    )

async def reset_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    CURRENT_PERIOD[user_id] = _this_month()
    clear_current_month(user_id)
    MODE[user_id] = "add"
    await update.message.reply_text("✅ Cleared this month's expenses. Add new items now.")

async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await start(update, context)

async def addexpense(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args or len(context.args) < 2:
        return await update.message.reply_text("Usage: /addexpense <amount> <category> [notes]")
    try:
        amount = float(context.args[0].replace(",", ""))
    except ValueError:
        return await update.message.reply_text("Amount must be a number.")
    category = context.args[1]
    notes = " ".join(context.args[2:]) if len(context.args) > 2 else ""
    user_id = update.effective_user.id
    add_expense_to_store(user_id, amount, category, notes)
    await update.message.reply_text(f"✅ Added ${amount:.2f} to '{category}'.")

async def addmany(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    /addmany\n1200 rent\n500 groceries\n200 gas truck
    or:
    /addmany  (then paste lines in the same message)
    """
    text = update.message.text
    block = text.split("\n", 1)[1] if "\n" in text else ""
    if not block.strip():
        return await update.message.reply_text(
            "Paste multiple lines after the command, e.g.:\n"
            "/addmany\n1200 rent\n500 groceries\n200 gas\n\n"
            "You can also separate with slashes: 1200 rent / 500 food / 200 insurance"
        )
    user_id = update.effective_user.id
    items, errors = parse_expense_block(block)
    for amt, cat, notes in items:
        add_expense_to_store(user_id, amt, cat, notes)
    msg = f"✅ Added {len(items)} item(s)."
    if errors:
        msg += f"\n⚠️ Skipped {len(errors)} line(s):\n- " + "\n- ".join(errors[:5])
        if len(errors) > 5:
            msg += f"\n(and {len(errors)-5} more...)"
    await update.message.reply_text(msg)

async def viewexpenses(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        user_id = update.effective_user.id
        items = get_current_month_expenses(user_id)
        if not items:
            return await update.message.reply_text("No expenses yet.")
        total = sum(x["amount"] for x in items)
        by_cat = {}
        for x in items:
            by_cat[x["category"]] = by_cat.get(x["category"], 0) + x["amount"]
        period = CURRENT_PERIOD[user_id]
        lines = [f"📊 {period} total: ${total:.2f}"]
        for k, v in sorted(by_cat.items(), key=lambda kv: -kv[1]):
            lines.append(f"- {k}: ${v:.2f}")
        await update.message.reply_text("\n".join(lines))
    except Exception as e:
        await update.message.reply_text(f"❌ view error: {e}")

async def generatebudget(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        user_id = update.effective_user.id
        items = get_current_month_expenses(user_id)
        if not items:
            return await update.message.reply_text("No data yet.")
        monthly_total = sum(x["amount"] for x in items)
        by_cat = {}
        for x in items:
            by_cat[x["category"]] = by_cat.get(x["category"], 0) + x["amount"]
        period = CURRENT_PERIOD[user_id]
        lines = [f"📅 Budget ({period}):", f"Total: ${monthly_total:.2f}", ""]
        for k, v in sorted(by_cat.items(), key=lambda kv: -kv[1]):
            lines.append(f"• {k}: ${v:.2f}")
        await update.message.reply_text("\n".join(lines))
    except Exception as e:
        await update.message.reply_text(f"❌ generate error: {e}")

# =============== Stripe Checkout ===============
async def unlockfull(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Create a $1 Stripe Checkout session and send the link."""
    user_id = update.effective_user.id
    period = CURRENT_PERIOD.get(user_id, _this_month())
    try:
        session = stripe.checkout.Session.create(
            mode="payment",
            success_url=f"{PUBLIC_BASE_URL}/stripe/success?u={user_id}&p={period}",
            cancel_url=f"{PUBLIC_BASE_URL}/stripe/cancel",
            line_items=[{
                "price_data": {
                    "currency": "usd",
                    "product_data": {"name": f"Budget Wizard unlock for {period}"},
                    "unit_amount": 100,  # $1.00 in cents
                },
                "quantity": 1
            }],
            metadata={"user_id": str(user_id), "period": period},
        )
        await update.message.reply_text(
            "Unlock the full report for **$1** via Stripe:\n"
            f"{session.url}\n\n"
            "After payment, I’ll unlock automatically."
        )
    except Exception as e:
        await update.message.reply_text(f"❌ Could not create payment link: {e}")

async def exportexcel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        user_id = update.effective_user.id
        items = get_current_month_expenses(user_id)
        if not items:
            return await update.message.reply_text("No expenses yet.")

        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"
        ws.append(["Period", "Timestamp", "Amount", "Category", "Notes"])

        if _is_unlocked(user_id):
            for x in items:
                ws.append([CURRENT_PERIOD[user_id], x["ts"].strftime("%Y-%m-%d %H:%M:%S"),
                           float(x["amount"]), x["category"], x["notes"]])
            fname, caption = "budget_full.xlsx", "✅ Full export unlocked."
        else:
            n = max(1, len(items)//2)
            for x in items[:n]:
                ws.append([CURRENT_PERIOD[user_id], x["ts"].strftime("%Y-%m-%d %H:%M:%S"),
                           float(x["amount"]), x["category"], x["notes"]])
            ws.insert_rows(1)
            ws["A1"] = "SAMPLE — Unlock full for $1 via /unlock"
            ws.merge_cells("A1:E1")
            ws["A1"].font = Font(bold=True)
            fname, caption = "budget_sample.xlsx", "This is a SAMPLE. Use /unlock to get the full report."

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        await update.message.reply_document(InputFile(bio, fname), caption=caption)
    except Exception as e:
        await update.message.reply_text(f"❌ export error: {e}")

# =============== Fallback Text ===============
async def fallback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text_raw = update.message.text or ""
    text = text_raw.strip().lower()
    user_id = update.effective_user.id

    # quick commands
    if text in ["reset", "new month", "start new month", "clear"]:
        clear_current_month(user_id)
        MODE[user_id] = "add"
        return await update.message.reply_text("✅ New month started. Paste expenses or type **done**.")
    if text in ["view", "summary"]:
        return await viewexpenses(update, context)
    if text in ["generate", "budget"]:
        return await generatebudget(update, context)
    if text in ["export", "excel"]:
        return await exportexcel(update, context)
    if text in ["unlock", "buy", "pay"]:
        return await unlockfull(update, context)

    # "add ..." with block support (slashes or newlines)
    if text.startswith("add "):
        content = update.message.text[4:]
        items, errors = parse_expense_block(content)
        if items:
            for amt, cat, notes in items:
                add_expense_to_store(user_id, amt, cat, notes)
            msg = f"✅ Added {len(items)} item(s)."
            if errors:
                msg += f"\n⚠️ Skipped {len(errors)} line(s):\n- " + "\n- ".join(errors[:5])
                if len(errors) > 5:
                    msg += f"\n(and {len(errors)-5} more...)"
            return await update.message.reply_text(msg)
        return await update.message.reply_text("Usage:\nadd 1200 rent / 500 food / 200 insurance")

    # Add-mode: handle multi-line and/or slashes in one go
    if MODE.get(user_id) == "add":
        block = text_raw.strip()
        items, errors = parse_expense_block(block)
        if items:
            for amt, cat, notes in items:
                add_expense_to_store(user_id, amt, cat, notes)
            msg = f"✅ Added {len(items)} item(s)."
            if errors:
                msg += f"\n⚠️ Skipped {len(errors)} line(s):\n- " + "\n- ".join(errors[:5])
                if len(errors) > 5:
                    msg += f"\n(and {len(errors)-5} more...)"
            return await update.message.reply_text(msg)

        return await update.message.reply_text(
            "Send expenses like:\n"
            "• 1200 rent\n• 500 food\n• 200 insurance\n"
            "or in one line: 1200 rent / 500 food / 200 insurance"
        )

    if text in ["done", "stop", "finish"]:
        MODE[user_id] = None
        return await update.message.reply_text("Okay. You can type **view**, **generate**, or **export** anytime.")

    return await update.message.reply_text(
        "Try: **hi** to start, `add 1200 rent`, paste multiple lines or use slashes, **view**, **generate**, **export**, **unlock**."
    )

# =============== Flask (Stripe Webhook) ===============
flask_app = Flask(__name__)

@flask_app.route("/stripe/webhook", methods=["POST"])
def stripe_webhook():
    payload = request.data
    sig_header = request.headers.get("Stripe-Signature", "")
    try:
        event = stripe.Webhook.construct_event(
            payload=payload, sig_header=sig_header, secret=STRIPE_WEBHOOK_SECRET
        )
    except Exception as e:
        return jsonify({"ok": False, "error": f"signature_error: {e}"}), 400

    if event["type"] == "checkout.session.completed":
        session = event["data"]["object"]
        if session.get("payment_status") == "paid":
            meta = session.get("metadata") or {}
            uid = meta.get("user_id")
            period = meta.get("period")
            if uid and period:
                try:
                    uid_int = int(uid)
                    _mark_unlocked(uid_int, period)
                    CURRENT_PERIOD[uid_int] = period
                    print(f"✅ Stripe verified & unlocked — user {uid_int}, period {period}")
                except ValueError:
                    pass
    return jsonify({"ok": True})

@flask_app.route("/stripe/success")
def stripe_success():
    return "Payment received! Return to Telegram and type 'export' to download your full report."

@flask_app.route("/stripe/cancel")
def stripe_cancel():
    return "Payment canceled."

def _run_flask():
    port = int(os.getenv("PORT", "8080"))
    flask_app.run(host="0.0.0.0", port=port, debug=False)

# =============== Main ===============
def main():
    if not TOKEN:
        raise SystemExit("BOT_TOKEN not set.")
    if not all([STRIPE_PUBLISHABLE_KEY, STRIPE_SECRET_KEY, STRIPE_WEBHOOK_SECRET, PUBLIC_BASE_URL]):
        print("⚠️ Missing Stripe or PUBLIC_BASE_URL envs — payments won’t work until set.")

    # Start webhook server
    threading.Thread(target=_run_flask, daemon=True).start()

    app = Application.builder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_cmd))
    app.add_handler(CommandHandler("reset", reset_cmd))
    app.add_handler(CommandHandler("addexpense", addexpense))
    app.add_handler(CommandHandler("addmany", addmany))
    app.add_handler(CommandHandler("viewexpenses", viewexpenses))
    app.add_handler(CommandHandler("generatebudget", generatebudget))
    app.add_handler(CommandHandler("exportexcel", exportexcel))
    app.add_handler(CommandHandler("unlockfull", unlockfull))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, fallback))
    app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == "__main__":
    main()
